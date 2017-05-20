from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

def write_xl(filename, sheet_name, ini_cell, data, labels):
	try:
		wb = load_workbook(filename)
	except:
		wb = Workbook()

	if wb.sheetnames.count(sheet_name) > 0:
		ws = wb[sheet_name]
	else:
		ws = wb.create_sheet(title=sheet_name)
	for j in range(len(labels)):
		cell = ws.cell(
			row = ws[ini_cell].row,
			column = column_index_from_string(ws[ini_cell].column) + j
		)
		cell.value = labels[j]
	for i in range(len(data)):
		dic = data[i]
		for j in range(len(labels)):
			cell = ws.cell(
				row = ws[ini_cell].row + 1 + i,
				column = column_index_from_string(ws[ini_cell].column) + j
			)
			cell.value = dic[labels[j]]
	wb.save(filename)
def range_xl(filename, sheet_name, cell_range):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    cells = ws[cell_range]
    return cells
def data_build(mat):
    keys = mat[0]
    coll = []
    for i in range(1, len(mat)):
        dic = {}
        for j in range(0, len(keys)):
            k = keys[j]
            if k != None and k != "":
                dic[k] = mat[i][j]
        coll.append(dic)
    return coll
def matrix_xl(cells):
    mat = []
    for row_xl in cells:
        row = []
        for cell in row_xl:
            row.append(cell.value)
        mat.append(row)
    return mat
def load_xl(filename, sheet_name, cell_range):
	cells = range_xl(filename, sheet_name, cell_range)
	mat = matrix_xl(cells)
	return data_build(mat)
def data_map(data, fn_map):
	aux = []
	for dic in data:
		x = fn_map(dic)
		if x != None:
			aux.append(x)
	return aux
def filtro(cliente):
    mes = cliente['Fecha'].split('-')
    if int(mes[1]) == 1:
        return '%s,%d' % (mes[1], cliente['Consumo'])
