# -*- coding: utf-8 -*-

'''
Diplomado Python Programming:
https://github.com/enriquemore587/Python-2017

ejemplos de diferentes codigos en lenguajes de programacion distinta.
https://gist.github.com/enriquemore587

openoyxl:
Es una libreria la cual nos permite trabajar con archivos de Excel en sus diferendes formatos.

Referencia:
https://openpyxl.readthedocs.io/en/default/

'''
#PASOS A SEGUIR PARA LA LECTURA DE UN DOCUMENTO DE EXCEL CON EXTENCION ".xlsx"

from openpyxl import load_workbook
'''
	DEBEMOS DE TENER EN CUENTA QUE UN DOCUMENTO DE EXCEL ES UN LIBRO EL CUAL TIENE HOJAS DENTRO DE EL MISMO 
	COMO CUALQUIER OTRO LIBRO QUE SEA FISICO
'''
###################################################### PRIMER PASO.
#TENEMOS QUE CARGAR A NUESTRO PROGRAMA EL LIBRO(workbook)
wb = load_workbook("Libro1.xlsx")




###################################################### SEGUNDO PASO.
# TENEMOS QUE ADQUIRIR LA HOJA (worksheet) QUE QUEREREMOS TRABAJAR  [TENER EN CUENTA QUE PUEDE TENER MAS DE UNA HOJA]
#		SE PUEDE MENSIONAR LA HOJA A TRABAJAR O BIEN CON EL METODO ACTIVE
# NOS DEVUELVE O CARGAR LA ULTIMA HOJA CON LA QUE SE TRABAJO

#ws = wb["Hoja 1"]
ws = wb.active


###################################################### TERCER PASO
#OPTENER EL RANGO DE LAS CELDAS A TRABAJAR
cells = ws["A1:B23"]

###################################################### CUARTO PASO
# PROCESAMIENTO DE LOS DATOS.
for row in cells:
    for cell in row:
        print cell.value

    print "-" * 30